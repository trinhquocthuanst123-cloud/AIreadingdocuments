import os
import pandas as pd
import json
from datetime import datetime

from openpyxl import Workbook
from openpyxl import load_workbook

from config import *
from models import *



BASE_DIR = os.path.dirname(os.path.abspath(__file__))

OUTPUT_DIR = os.path.join(BASE_DIR, OUTPUT_FOLDER)

os.makedirs(OUTPUT_DIR, exist_ok=True)

EXCEL_PATH = os.path.join(
    OUTPUT_DIR,
    EXCEL_FILE
)
SHEETS = [

    "Dashboard",

    "Purchase Order",

    "Sales Contract",

    "Commercial Invoice",

    "Packing List",

    "Bill of Lading",

    "Sea Waybill",

    "Air Waybill",

    "Booking Confirmation",

    "Delivery Order",

    "Arrival Notice",

    "Certificate of Origin",

    "Insurance Certificate",

    "Manifest",

    "Health Certificate",

    "Phytosanitary Certificate",

    "Radiation Certificate",

    "Quality & Quantity Certificate"

]
def create_workbook():

    if os.path.exists(EXCEL_PATH):
        return

    wb = Workbook()
    wb.remove(wb.active)

    for sheet in SHEETS:

        ws = wb.create_sheet(sheet)

        if sheet == "Dashboard":

            headers = [
                "Processed At",
                "Document Type",
                "Document Number",
                "Buyer",
                "Seller",
                "PO",
                "Invoice",
                "BL",
                "Container"
            ]

        else:

            headers = DOCUMENT_FIELDS.get(sheet, [])

        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col).value = header

    wb.save(EXCEL_PATH)
def get_sheet(sheet_name):

    create_workbook()

    wb = load_workbook(EXCEL_PATH)

    ws = wb[sheet_name]

    return wb, ws
def dashboard_row(data):

    return {

        "Processed At": datetime.now().strftime(DATE_FORMAT),

        "Document Type": data.get("document_type"),

        "Document Number": data.get("document_number"),

        "Buyer": data.get("buyer"),

        "Seller": data.get("seller"),

        "PO": data.get("po_number"),

        "Invoice": data.get("invoice_number"),

        "BL": data.get("bl_number"),

        "Container": data.get("container_number")

    }
def document_row(data):

    doc = data.get(
        "document_type",
        ""
    )

    fields = DOCUMENT_FIELDS.get(doc)

    if fields is None:
        fields = ALL_FIELDS

    row = {}

    for field in fields:

        row[field] = data.get(field)

    return row
def excel_value(value):

    if value is None:
        return ""

    if isinstance(value, list):
        return ", ".join(str(v) for v in value)

    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)

    return value
def append_row(sheet_name, row):

    create_workbook()

    wb = load_workbook(EXCEL_PATH)

    ws = wb[sheet_name]

    headers = []

    if ws.max_row > 0:
        headers = [cell.value for cell in ws[1] if cell.value]

    if not headers:
        headers = list(row.keys())

        for i, h in enumerate(headers, start=1):
            ws.cell(row=1, column=i).value = h
    print("Sheet:", sheet_name)
    print("Headers:", headers)
    print("Row:", row)
    header_map = {
        h: i + 1
        for i, h in enumerate(headers)
    }

    new_row = ws.max_row + 1

    for key, value in row.items():
        print("KEY =", key)
        print("VALUE =", value)
        print("IN HEADER =", key in header_map)
        print("KEY =", key)

        print("HEADER MAP =", header_map)

        if key not in header_map:
            print(">>> Missing:", key)
            continue

        ws.cell(
            row=new_row,
            column=header_map[key]
        ).value = excel_value(value)

    wb.save(EXCEL_PATH)

def write_dashboard(data):

    row = dashboard_row(data)

    append_row(
        "Dashboard",
        row
    )
def write_document(data):

    doc = data.get(
        "document_type",
        "Dashboard"
    )
    print("WRITE DOCUMENT")
    print("Document Type =", doc)
    print("Sheet Exists =", doc in SHEETS)
    if doc not in SHEETS:

        doc = "Dashboard"

    row = document_row(data)
    print("ROW =", row)
    append_row(
        doc,
        row
    )

def export_excel(data):

    # 1. Ghi Dashboard
    write_dashboard(data)

    # 2. Ghi đúng sheet theo loại chứng từ
    write_document(data)

    return EXCEL_PATH